from flask import Blueprint, render_template, request, redirect, url_for
from flask_login import login_required, current_user
from urllib.parse import quote
from openpyxl import load_workbook
import requests

from app import db
from app.models.lead import Lead
from app.models.pipeline import Pipeline
from app.models.interacao import Interacao
from app.models.empresa import Empresa
from app.models.usuario import Usuario
from app.services.ia_sdr import responder_lead, registrar_interacao_ia


leads_bp = Blueprint("leads", __name__, url_prefix="/leads")


def garantir_empresa_usuario():
    if not current_user.empresa_id:
        empresa = Empresa(nome="Empresa CRM360", plano="teste")
        db.session.add(empresa)
        db.session.flush()

        current_user.empresa_id = empresa.id

        leads_sem_empresa = Lead.query.filter_by(empresa_id=None).all()
        for lead in leads_sem_empresa:
            lead.empresa_id = empresa.id
            lead.usuario_id = current_user.id

        db.session.commit()


def criar_pipeline_padrao():
    garantir_empresa_usuario()

    etapas = [
        ("Novo Lead", 1),
        ("Contato Feito", 2),
        ("Proposta", 3),
        ("Negociação", 4),
        ("Fechado", 5),
        ("Perdido", 6),
    ]

    for nome, ordem in etapas:
        existe = Pipeline.query.filter_by(
            nome=nome,
            empresa_id=current_user.empresa_id
        ).first()

        if not existe:
            etapa = Pipeline(
                nome=nome,
                ordem=ordem,
                empresa_id=current_user.empresa_id
            )
            db.session.add(etapa)

    db.session.commit()


def registrar_historico_whatsapp(lead, tipo, descricao):
    interacao = Interacao(
        lead_id=lead.id,
        empresa_id=current_user.empresa_id,
        usuario_id=current_user.id,
        tipo=tipo,
        descricao=descricao
    )

    db.session.add(interacao)
    db.session.commit()


def limpar_texto(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def limpar_telefone(valor):
    telefone = limpar_texto(valor)
    return ''.join(filter(str.isdigit, telefone))


@leads_bp.route("/")
@login_required
def lista():
    criar_pipeline_padrao()

    leads = Lead.query.filter_by(
        empresa_id=current_user.empresa_id
    ).all()

    return render_template("leads.html", leads=leads)


@leads_bp.route("/novo", methods=["GET", "POST"])
@login_required
def novo():
    criar_pipeline_padrao()

    primeira_etapa = Pipeline.query.filter_by(
        empresa_id=current_user.empresa_id
    ).order_by(Pipeline.ordem.asc()).first()

    usuarios = Usuario.query.filter_by(
        empresa_id=current_user.empresa_id
    ).order_by(Usuario.nome.asc()).all()

    if request.method == "POST":
        usuario_id = request.form.get("usuario_id") or current_user.id
        instagram = request.form.get("instagram", "").strip()
        plano = request.form.get("plano", "").strip()
        status = request.form.get("status", "aberto").strip()
        valor = request.form.get("valor") or 0

        lead = Lead(
            nome=request.form["nome"],
            telefone=request.form["telefone"],
            instagram=instagram,
            origem=request.form["origem"],
            produto_interesse=request.form["produto"],
            plano=plano,
            valor=float(valor),
            status=status,
            pipeline_id=primeira_etapa.id if primeira_etapa else None,
            empresa_id=current_user.empresa_id,
            usuario_id=usuario_id
        )
        db.session.add(lead)
        db.session.commit()

        return redirect(url_for("leads.lista"))

    return render_template("lead_form.html", usuarios=usuarios)


@leads_bp.route("/importar", methods=["GET", "POST"])
@login_required
def importar():
    criar_pipeline_padrao()

    primeira_etapa = Pipeline.query.filter_by(
        empresa_id=current_user.empresa_id
    ).order_by(Pipeline.ordem.asc()).first()

    usuarios = Usuario.query.filter_by(
        empresa_id=current_user.empresa_id
    ).order_by(Usuario.nome.asc()).all()

    if request.method == "POST":
        arquivo = request.files.get("arquivo")
        usuario_id = request.form.get("usuario_id") or current_user.id

        if not arquivo or arquivo.filename == "":
            return render_template(
                "lead_importar.html",
                usuarios=usuarios,
                erro="Selecione um arquivo Excel para importar."
            )

        if not arquivo.filename.lower().endswith((".xlsx", ".xlsm")):
            return render_template(
                "lead_importar.html",
                usuarios=usuarios,
                erro="Envie um arquivo no formato .xlsx ou .xlsm."
            )

        workbook = load_workbook(arquivo, data_only=True)
        planilha = workbook.active

        importados = 0
        ignorados = 0

        for linha in planilha.iter_rows(min_row=2, values_only=True):
            nome = limpar_texto(linha[0] if len(linha) > 0 else "")
            telefone = limpar_telefone(linha[1] if len(linha) > 1 else "")
            email = limpar_texto(linha[2] if len(linha) > 2 else "")
            origem = limpar_texto(linha[3] if len(linha) > 3 else "")
            produto = limpar_texto(linha[4] if len(linha) > 4 else "")
            instagram = limpar_texto(linha[5] if len(linha) > 5 else "")

            if not nome and not telefone:
                ignorados += 1
                continue

            existe = None
            if telefone:
                existe = Lead.query.filter_by(
                    empresa_id=current_user.empresa_id,
                    telefone=telefone
                ).first()

            if existe:
                ignorados += 1
                continue

            lead = Lead(
                nome=nome or "Lead sem nome",
                telefone=telefone,
                email=email,
                instagram=instagram,
                origem=origem or "Importação Excel",
                produto_interesse=produto,
                pipeline_id=primeira_etapa.id if primeira_etapa else None,
                empresa_id=current_user.empresa_id,
                usuario_id=usuario_id
            )

            db.session.add(lead)
            importados += 1

        db.session.commit()

        return render_template(
            "lead_importar.html",
            usuarios=usuarios,
            sucesso=f"Importação concluída: {importados} lead(s) importado(s), {ignorados} ignorado(s)."
        )

    return render_template("lead_importar.html", usuarios=usuarios)


@leads_bp.route("/<int:lead_id>")
@login_required
def detalhe(lead_id):
    criar_pipeline_padrao()

    lead = Lead.query.filter_by(
        id=lead_id,
        empresa_id=current_user.empresa_id
    ).first_or_404()

    interacoes = Interacao.query.filter_by(
        lead_id=lead.id,
        empresa_id=current_user.empresa_id
    ).order_by(Interacao.criado_em.desc()).all()

    return render_template(
        "lead_detalhe.html",
        lead=lead,
        interacoes=interacoes
    )


@leads_bp.route("/<int:lead_id>/interacao", methods=["POST"])
@login_required
def nova_interacao(lead_id):
    criar_pipeline_padrao()

    lead = Lead.query.filter_by(
        id=lead_id,
        empresa_id=current_user.empresa_id
    ).first_or_404()

    tipo = request.form["tipo"]
    descricao = request.form["descricao"]

    interacao = Interacao(
        lead_id=lead.id,
        empresa_id=current_user.empresa_id,
        usuario_id=current_user.id,
        tipo=tipo,
        descricao=descricao
    )

    db.session.add(interacao)
    db.session.commit()

    return redirect(url_for("leads.detalhe", lead_id=lead.id))


@leads_bp.route("/<int:lead_id>/whatsapp-manual", methods=["POST"])
@login_required
def whatsapp_manual(lead_id):
    criar_pipeline_padrao()

    lead = Lead.query.filter_by(
        id=lead_id,
        empresa_id=current_user.empresa_id
    ).first_or_404()

    mensagem = request.form.get("mensagem_whatsapp", "").strip()

    if not mensagem:
        mensagem = f"Olá {lead.nome}, tudo bem? Estou entrando em contato pelo CRM360."

    registrar_historico_whatsapp(
        lead,
        "WhatsApp manual",
        f"Mensagem preparada/enviada manualmente: {mensagem}"
    )

    telefone = lead.telefone or ""
    telefone = ''.join(filter(str.isdigit, telefone))

    if not telefone:
        registrar_historico_whatsapp(
            lead,
            "WhatsApp erro",
            "Não foi possível abrir o WhatsApp: telefone inválido ou vazio."
        )
        return redirect(url_for("leads.detalhe", lead_id=lead.id))

    link = f"https://wa.me/55{telefone}?text={quote(mensagem)}"

    return redirect(link)


@leads_bp.route("/<int:lead_id>/ia", methods=["POST"])
@login_required
def ia_responder(lead_id):
    criar_pipeline_padrao()

    lead = Lead.query.filter_by(
        id=lead_id,
        empresa_id=current_user.empresa_id
    ).first_or_404()

    mensagem = request.form["mensagem"]

    resposta = responder_lead(lead, mensagem)

    registrar_interacao_ia(lead, mensagem, resposta)

    return redirect(url_for("leads.detalhe", lead_id=lead.id))


@leads_bp.route("/<int:lead_id>/whatsapp-api", methods=["POST"])
@login_required
def whatsapp_api(lead_id):
    criar_pipeline_padrao()

    lead = Lead.query.filter_by(
        id=lead_id,
        empresa_id=current_user.empresa_id
    ).first_or_404()

    empresa = Empresa.query.get(current_user.empresa_id)

    mensagem = request.form.get("mensagem_whatsapp", "").strip()

    if not mensagem:
        mensagem = f"Olá {lead.nome}, tudo bem? Estou entrando em contato pelo CRM360."

    if not empresa:
        registrar_historico_whatsapp(
            lead,
            "WhatsApp erro",
            "Erro no envio via API: empresa não encontrada."
        )
        return redirect(url_for("leads.detalhe", lead_id=lead.id))

    if not empresa.whatsapp_ativo:
        registrar_historico_whatsapp(
            lead,
            "WhatsApp API desativada",
            "Tentativa de envio via API, mas a API do WhatsApp está desativada nas configurações."
        )
        return redirect(url_for("leads.detalhe", lead_id=lead.id))

    if not empresa.whatsapp_token or not empresa.whatsapp_phone_number_id:
        registrar_historico_whatsapp(
            lead,
            "WhatsApp erro",
            "Erro no envio via API: Token ou Phone Number ID não configurado."
        )
        return redirect(url_for("leads.detalhe", lead_id=lead.id))

    telefone = lead.telefone or ""
    telefone = ''.join(filter(str.isdigit, telefone))

    if not telefone:
        registrar_historico_whatsapp(
            lead,
            "WhatsApp erro",
            "Erro no envio via API: telefone do lead inválido ou vazio."
        )
        return redirect(url_for("leads.detalhe", lead_id=lead.id))

    url = f"https://graph.facebook.com/v18.0/{empresa.whatsapp_phone_number_id}/messages"

    headers = {
        "Authorization": f"Bearer {empresa.whatsapp_token}",
        "Content-Type": "application/json"
    }

    payload = {
        "messaging_product": "whatsapp",
        "to": f"55{telefone}",
        "type": "text",
        "text": {
            "body": mensagem
        }
    }

    try:
        response = requests.post(url, headers=headers, json=payload, timeout=20)

        if response.status_code in [200, 201]:
            registrar_historico_whatsapp(
                lead,
                "WhatsApp API enviado",
                f"Mensagem enviada com sucesso via API: {mensagem}"
            )
        else:
            registrar_historico_whatsapp(
                lead,
                "WhatsApp API erro",
                f"Erro ao enviar via API. Status: {response.status_code}. Retorno: {response.text}"
            )

    except requests.exceptions.RequestException as erro:
        registrar_historico_whatsapp(
            lead,
            "WhatsApp API erro",
            f"Falha de conexão com a API da Meta: {erro}"
        )

    return redirect(url_for("leads.detalhe", lead_id=lead.id))


@leads_bp.route("/<int:lead_id>/ligar")
@login_required
def ligar(lead_id):
    criar_pipeline_padrao()

    lead = Lead.query.filter_by(
        id=lead_id,
        empresa_id=current_user.empresa_id
    ).first_or_404()

    telefone = lead.telefone or ""
    telefone = ''.join(filter(str.isdigit, telefone))

    registrar_historico_whatsapp(
        lead,
        "Ligação iniciada",
        f"Ligação iniciada para o número {telefone}"
    )

    if not telefone:
        registrar_historico_whatsapp(
            lead,
            "Ligação erro",
            "Tentativa de ligação sem telefone válido."
        )
        return redirect(url_for("leads.detalhe", lead_id=lead.id))

    return redirect(f"tel:{telefone}")

@leads_bp.route("/<int:lead_id>/tags", methods=["POST"])
@login_required
def atualizar_tags(lead_id):

    lead = Lead.query.filter_by(
        id=lead_id,
        empresa_id=current_user.empresa_id
    ).first_or_404()

    tags = request.form.getlist("tags")

    lead.tags = ",".join(tags)

    db.session.commit()

    registrar_historico_whatsapp(
        lead,
        "Tags atualizadas",
        f"Tags definidas: {lead.tags}"
    )

    return redirect(
        url_for(
            "leads.detalhe",
            lead_id=lead.id
        )
    )