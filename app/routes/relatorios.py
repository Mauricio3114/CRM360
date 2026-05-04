from flask import Blueprint, render_template, request, send_file
from flask_login import login_required, current_user
from io import BytesIO
from datetime import datetime

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from openpyxl import Workbook

from app.models.lead import Lead
from app.models.usuario import Usuario
from app.models.lancamento_financeiro import LancamentoFinanceiro

relatorios_bp = Blueprint("relatorios", __name__, url_prefix="/relatorios")


def moeda(valor):
    return f"R$ {valor:.2f}"


def usuario_nome_por_id(usuario_id):
    usuario = Usuario.query.get(usuario_id)
    return usuario.nome if usuario else "Sem vendedor"


def filtrar_periodo(query):
    data_inicio = request.args.get("data_inicio")
    data_fim = request.args.get("data_fim")

    if data_inicio:
        inicio = datetime.strptime(data_inicio, "%Y-%m-%d")
        query = query.filter(LancamentoFinanceiro.data_lancamento >= inicio)

    if data_fim:
        fim = datetime.strptime(data_fim, "%Y-%m-%d")
        query = query.filter(LancamentoFinanceiro.data_lancamento <= fim)

    return query


def filtrar_leads():
    vendedor_id = request.args.get("vendedor_id")

    query = Lead.query.filter_by(empresa_id=current_user.empresa_id)

    if vendedor_id:
        query = query.filter_by(usuario_id=vendedor_id)

    return query.all()


def filtrar_vendas():
    vendedor_id = request.args.get("vendedor_id")

    query = LancamentoFinanceiro.query.filter_by(
        empresa_id=current_user.empresa_id,
        tipo="entrada"
    )

    query = filtrar_periodo(query)
    vendas = query.all()

    if vendedor_id:
        vendas = [
            venda for venda in vendas
            if venda.lead and str(venda.lead.usuario_id) == str(vendedor_id)
        ]

    return vendas


def filtrar_financeiro():
    tipo = request.args.get("tipo")

    query = LancamentoFinanceiro.query.filter_by(
        empresa_id=current_user.empresa_id
    )

    query = filtrar_periodo(query)

    if tipo:
        query = query.filter_by(tipo=tipo)

    return query.order_by(LancamentoFinanceiro.data_lancamento.desc()).all()


def montar_comissoes():
    vendedor_id = request.args.get("vendedor_id")
    vendas = filtrar_vendas()

    usuarios = Usuario.query.filter_by(
        empresa_id=current_user.empresa_id
    ).order_by(Usuario.nome.asc()).all()

    if vendedor_id:
        usuarios = [u for u in usuarios if str(u.id) == str(vendedor_id)]

    linhas = []

    for usuario in usuarios:
        vendas_usuario = [
            v for v in vendas
            if v.lead and v.lead.usuario_id == usuario.id
        ]

        total_vendas = sum(v.valor for v in vendas_usuario)
        percentual = usuario.percentual_comissao or 0
        comissao = (total_vendas * percentual) / 100

        linhas.append({
            "usuario": usuario,
            "total_vendas": total_vendas,
            "percentual": percentual,
            "comissao": comissao,
            "qtd_vendas": len(vendas_usuario)
        })

    return linhas


@relatorios_bp.route("/")
@login_required
def index():
    usuarios = Usuario.query.filter_by(
        empresa_id=current_user.empresa_id
    ).order_by(Usuario.nome.asc()).all()

    return render_template("relatorios.html", usuarios=usuarios)


@relatorios_bp.route("/leads/pdf")
@login_required
def leads_pdf():
    leads = filtrar_leads()
    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=25, leftMargin=25, topMargin=25, bottomMargin=25)
    styles = getSampleStyleSheet()
    elementos = [Paragraph("Relatório de Leads - CRM360", styles["Title"]), Spacer(1, 12)]

    dados = [["Nome", "Telefone", "Email", "Origem", "Produto", "Vendedor"]]

    for lead in leads:
        dados.append([
            lead.nome or "-",
            lead.telefone or "-",
            lead.email or "-",
            lead.origem or "-",
            lead.produto_interesse or "-",
            usuario_nome_por_id(lead.usuario_id) if lead.usuario_id else "Sem vendedor"
        ])

    tabela = Table(dados, repeatRows=1)
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))

    elementos.append(tabela)
    doc.build(elementos)
    buffer.seek(0)

    return send_file(buffer, as_attachment=True, download_name="relatorio_leads_crm360.pdf", mimetype="application/pdf")


@relatorios_bp.route("/leads/excel")
@login_required
def leads_excel():
    leads = filtrar_leads()

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(["Nome", "Telefone", "Email", "Origem", "Produto", "Vendedor"])

    for lead in leads:
        ws.append([
            lead.nome or "",
            lead.telefone or "",
            lead.email or "",
            lead.origem or "",
            lead.produto_interesse or "",
            usuario_nome_por_id(lead.usuario_id) if lead.usuario_id else "Sem vendedor"
        ])

    for coluna in ws.columns:
        letra = coluna[0].column_letter
        ws.column_dimensions[letra].width = max(len(str(c.value)) if c.value else 0 for c in coluna) + 4

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(buffer, as_attachment=True, download_name="relatorio_leads_crm360.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@relatorios_bp.route("/vendas/pdf")
@login_required
def vendas_pdf():
    vendas = filtrar_vendas()
    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=25, leftMargin=25, topMargin=25, bottomMargin=25)
    styles = getSampleStyleSheet()
    elementos = [Paragraph("Relatório de Vendas - CRM360", styles["Title"]), Spacer(1, 12)]

    total = sum(venda.valor for venda in vendas)
    dados = [["Data", "Descrição", "Valor", "Lead", "Vendedor"]]

    for venda in vendas:
        lead_nome = venda.lead.nome if venda.lead else "-"
        vendedor = usuario_nome_por_id(venda.lead.usuario_id) if venda.lead and venda.lead.usuario_id else "Sem vendedor"

        dados.append([
            venda.data_lancamento.strftime("%d/%m/%Y") if venda.data_lancamento else "-",
            venda.descricao or "-",
            moeda(venda.valor),
            lead_nome,
            vendedor
        ])

    dados.append(["", "TOTAL", moeda(total), "", ""])

    tabela = Table(dados, repeatRows=1)
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#dcfce7")),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))

    elementos.append(tabela)
    doc.build(elementos)
    buffer.seek(0)

    return send_file(buffer, as_attachment=True, download_name="relatorio_vendas_crm360.pdf", mimetype="application/pdf")


@relatorios_bp.route("/vendas/excel")
@login_required
def vendas_excel():
    vendas = filtrar_vendas()

    wb = Workbook()
    ws = wb.active
    ws.title = "Vendas"
    ws.append(["Data", "Descrição", "Valor", "Lead", "Vendedor"])

    total = 0

    for venda in vendas:
        total += venda.valor
        lead_nome = venda.lead.nome if venda.lead else ""
        vendedor = usuario_nome_por_id(venda.lead.usuario_id) if venda.lead and venda.lead.usuario_id else "Sem vendedor"

        ws.append([
            venda.data_lancamento.strftime("%d/%m/%Y") if venda.data_lancamento else "",
            venda.descricao or "",
            venda.valor,
            lead_nome,
            vendedor
        ])

    ws.append(["", "TOTAL", total, "", ""])

    for coluna in ws.columns:
        letra = coluna[0].column_letter
        ws.column_dimensions[letra].width = max(len(str(c.value)) if c.value else 0 for c in coluna) + 4

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(buffer, as_attachment=True, download_name="relatorio_vendas_crm360.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@relatorios_bp.route("/comissoes/pdf")
@login_required
def comissoes_pdf():
    linhas = montar_comissoes()
    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=25, leftMargin=25, topMargin=25, bottomMargin=25)
    styles = getSampleStyleSheet()
    elementos = [Paragraph("Relatório de Comissões - CRM360", styles["Title"]), Spacer(1, 12)]

    total_comissao = sum(item["comissao"] for item in linhas)

    dados = [["Vendedor", "Qtd. Vendas", "Total vendido", "% Comissão", "Comissão"]]

    for item in linhas:
        dados.append([
            item["usuario"].nome,
            item["qtd_vendas"],
            moeda(item["total_vendas"]),
            f'{item["percentual"]:.2f}%',
            moeda(item["comissao"])
        ])

    dados.append(["TOTAL", "", "", "", moeda(total_comissao)])

    tabela = Table(dados, repeatRows=1)
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#dcfce7")),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))

    elementos.append(tabela)
    doc.build(elementos)
    buffer.seek(0)

    return send_file(buffer, as_attachment=True, download_name="relatorio_comissoes_crm360.pdf", mimetype="application/pdf")


@relatorios_bp.route("/comissoes/excel")
@login_required
def comissoes_excel():
    linhas = montar_comissoes()

    wb = Workbook()
    ws = wb.active
    ws.title = "Comissões"
    ws.append(["Vendedor", "Qtd. Vendas", "Total vendido", "% Comissão", "Comissão"])

    total_comissao = 0

    for item in linhas:
        total_comissao += item["comissao"]
        ws.append([
            item["usuario"].nome,
            item["qtd_vendas"],
            item["total_vendas"],
            item["percentual"],
            item["comissao"]
        ])

    ws.append(["TOTAL", "", "", "", total_comissao])

    for coluna in ws.columns:
        letra = coluna[0].column_letter
        ws.column_dimensions[letra].width = max(len(str(c.value)) if c.value else 0 for c in coluna) + 4

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(buffer, as_attachment=True, download_name="relatorio_comissoes_crm360.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@relatorios_bp.route("/financeiro/pdf")
@login_required
def financeiro_pdf():
    lancamentos = filtrar_financeiro()
    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=25, leftMargin=25, topMargin=25, bottomMargin=25)
    styles = getSampleStyleSheet()
    elementos = [Paragraph("Relatório Financeiro - CRM360", styles["Title"]), Spacer(1, 12)]

    total_entradas = sum(l.valor for l in lancamentos if l.tipo == "entrada")
    total_saidas = sum(l.valor for l in lancamentos if l.tipo == "saida")
    lucro = total_entradas - total_saidas

    dados = [["Data", "Descrição", "Tipo", "Categoria", "Lead", "Valor"]]

    for item in lancamentos:
        dados.append([
            item.data_lancamento.strftime("%d/%m/%Y") if item.data_lancamento else "-",
            item.descricao or "-",
            item.tipo or "-",
            item.categoria or "-",
            item.lead.nome if item.lead else "-",
            moeda(item.valor)
        ])

    dados.append(["", "ENTRADAS", "", "", "", moeda(total_entradas)])
    dados.append(["", "SAÍDAS", "", "", "", moeda(total_saidas)])
    dados.append(["", "LUCRO", "", "", "", moeda(lucro)])

    tabela = Table(dados, repeatRows=1)
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -3), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -3), (-1, -1), colors.HexColor("#e0f2fe")),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))

    elementos.append(tabela)
    doc.build(elementos)
    buffer.seek(0)

    return send_file(buffer, as_attachment=True, download_name="relatorio_financeiro_crm360.pdf", mimetype="application/pdf")


@relatorios_bp.route("/financeiro/excel")
@login_required
def financeiro_excel():
    lancamentos = filtrar_financeiro()

    wb = Workbook()
    ws = wb.active
    ws.title = "Financeiro"
    ws.append(["Data", "Descrição", "Tipo", "Categoria", "Lead", "Valor"])

    total_entradas = 0
    total_saidas = 0

    for item in lancamentos:
        if item.tipo == "entrada":
            total_entradas += item.valor
        elif item.tipo == "saida":
            total_saidas += item.valor

        ws.append([
            item.data_lancamento.strftime("%d/%m/%Y") if item.data_lancamento else "",
            item.descricao or "",
            item.tipo or "",
            item.categoria or "",
            item.lead.nome if item.lead else "",
            item.valor
        ])

    lucro = total_entradas - total_saidas

    ws.append(["", "ENTRADAS", "", "", "", total_entradas])
    ws.append(["", "SAÍDAS", "", "", "", total_saidas])
    ws.append(["", "LUCRO", "", "", "", lucro])

    for coluna in ws.columns:
        letra = coluna[0].column_letter
        ws.column_dimensions[letra].width = max(len(str(c.value)) if c.value else 0 for c in coluna) + 4

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(buffer, as_attachment=True, download_name="relatorio_financeiro_crm360.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")